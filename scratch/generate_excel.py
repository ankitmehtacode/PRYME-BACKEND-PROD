import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json

def create_excel():
    wb = openpyxl.Workbook()
    
    # -------------------------------------------------------------
    # 1. OVERVIEW & GUIDE SHEET
    # -------------------------------------------------------------
    ws_overview = wb.active
    ws_overview.title = "Overview & Guide"
    ws_overview.views.sheetView[0].showGridLines = True
    
    # Title
    ws_overview["A2"] = "PRYME Loan Eligibility Engine - Comprehensive Test Suite"
    ws_overview["A2"].font = Font(name="Segoe UI", size=18, bold=True, color="1F497D")
    
    ws_overview["A3"] = "Contains 105 standard and edge-case scenarios to verify engine correctness."
    ws_overview["A3"].font = Font(name="Segoe UI", size=11, italic=True)
    
    # Info Block
    info = {
        "A5": "Project:", "B5": "PRYME Backend Eligibility Engine Test Suite",
        "A6": "Test Cases Count:", "B6": "105 Unique Scenarios",
        "A7": "Target Endpoints:", "B7": "/api/eligibility/evaluate (POST request)",
        "A8": "Last Updated:", "B8": "June 17, 2026",
    }
    for cell_ref, text in info.items():
        ws_overview[cell_ref] = text
        if cell_ref.startswith("A"):
            ws_overview[cell_ref].font = Font(name="Segoe UI", size=10, bold=True, color="595959")
        else:
            ws_overview[cell_ref].font = Font(name="Segoe UI", size=10)
            
    # Guidelines
    ws_overview["A11"] = "Test Categories & Coverage:"
    ws_overview["A11"].font = Font(name="Segoe UI", size=12, bold=True, color="1F497D")
    
    categories = [
        "1. CIBIL Score Thresholds (TC-001 to TC-015): Verify minimum CIBIL limit, and ROI selection based on bands.",
        "2. Age Limit & Maturity Checks (TC-016 to TC-030): Verify minimum age, maximum age, and maturity calculation: Age + Tenure/12.",
        "3. Income Floor Checks (TC-031 to TC-045): Verify monthly salary limit validations.",
        "4. Employment & Program Routing (TC-046 to TC-060): Verify NIP, SEP, GST, Banking, Low LTV routing.",
        "5. Property Type & Negative Lists (TC-061 to TC-075): Verify residential, commercial, plots, and negative lists.",
        "6. Pin Code & Geo-Fencing (TC-076 to TC-090): Verify service area restriction (Indore 452xxx/453xxx).",
        "7. SpEL & Decimal Boundary Check (TC-091 to TC-105): Verify exact boundaries (e.g. ₹60,000.01) and 'No Limit' matching."
    ]
    
    for i, cat in enumerate(categories):
        cell_ref = f"A{13 + i}"
        ws_overview[cell_ref] = cat
        ws_overview[cell_ref].font = Font(name="Segoe UI", size=11)
        
    # -------------------------------------------------------------
    # 2. GENERATE 105 TEST CASES DATA
    # -------------------------------------------------------------
    scenarios = []
    
    # --- Category 1: CIBIL Score Thresholds (TC-001 to TC-015) ---
    cibil_cases = [
        (550, "HDFC Bank", "HOME_LOAN", "SALARIED", "REJECTED", "CIBIL 550 is below HDFC HL floor of 650."),
        (580, "L&T Finance", "HOME_LOAN", "SALARIED", "REJECTED", "CIBIL 580 is below L&T HL floor of 650."),
        (600, "ICICI Bank", "LOAN_AGAINST_PROPERTY", "SALARIED", "REJECTED", "CIBIL 600 is below ICICI LAP floor of 650."),
        (620, "Yes Bank", "HOME_LOAN", "PROFESSIONAL", "REJECTED", "CIBIL 620 is below Yes Bank HL floor of 650."),
        (640, "Bajaj Prime", "HOME_LOAN", "SELF_EMPLOYED", "REJECTED", "CIBIL 640 is below Bajaj HL floor of 650."),
        (650, "HDFC Bank", "HOME_LOAN", "SALARIED", "ELIGIBLE", "CIBIL 650 is exactly on the floor limit. Matches lowest ROI tier."),
        (660, "Bandhan Bank", "HOME_LOAN", "SALARIED", "ELIGIBLE", "CIBIL 660 is eligible for Bandhan Bank Home Loan."),
        (680, "L&T Finance", "HOME_LOAN", "SALARIED", "ELIGIBLE", "CIBIL 680 is eligible. Standard ROI applies."),
        (700, "ICICI Bank", "HOME_LOAN", "SALARIED", "ELIGIBLE", "CIBIL 700 is eligible. Maps to mid-range ROI matrix."),
        (720, "Bank of Baroda", "HOME_LOAN", "PROFESSIONAL", "ELIGIBLE", "CIBIL 720 matches professional lane requirements."),
        (740, "Yes Bank", "LOAN_AGAINST_PROPERTY", "SELF_EMPLOYED", "ELIGIBLE", "CIBIL 740 is eligible. High credit score profile."),
        (750, "SBI", "HOME_LOAN", "SALARIED", "ELIGIBLE", "CIBIL 750 is eligible. Prime SBI Home Loan ROI applies."),
        (780, "Aditya Birla Finance Limited", "LOAN_AGAINST_PROPERTY", "SELF_EMPLOYED", "ELIGIBLE", "CIBIL 780 matches prime tier. Best ROI and LTV conditions."),
        (800, "TATA Capital", "LOAN_AGAINST_PROPERTY", "SALARIED", "ELIGIBLE", "CIBIL 800 is eligible. Super prime applicant status."),
        (850, "Jio Finance", "HOME_LOAN", "PROFESSIONAL", "ELIGIBLE", "CIBIL 850 is eligible. Outstanding credit score profile.")
    ]
    for idx, (cibil, lender, loan_type, emp, dec, reason) in enumerate(cibil_cases, 1):
        scenarios.append({
            "id": f"TC-{idx:03d}",
            "category": "CIBIL Score Thresholds",
            "objective": f"Verify eligibility rules for CIBIL={cibil} with {lender}",
            "lender": lender,
            "product": loan_type,
            "employment": emp,
            "cibil": cibil,
            "income": "Net: ₹100,000",
            "amount": 3000000,
            "tenure": 240,
            "age": 35,
            "expected_ltv": "75% to 80%" if dec == "ELIGIBLE" else "N/A",
            "expected_foir": "65% to 80%" if dec == "ELIGIBLE" else "N/A",
            "decision": dec,
            "reason": reason,
            "payload_template": {
                "lenderId": 73 if "L&T" in lender else (200 if "ICICI" in lender else 105),
                "loanType": loan_type,
                "cibilScore": cibil,
                "applicantAge": 35,
                "employmentType": emp,
                "loanAmount": 3000000.00,
                "propertyValue": 4000000.00,
                "requestedTenureMonths": 240,
                "monthlyIncome": 100000.00,
                "existingEmiTotal": 0.00,
                "idempotencyKey": f"test-cibil-{idx:03d}",
                "pinCode": "452001",
                "incomeComputationInput": {"programName": "NIP"}
            }
        })
        
    # --- Category 2: Age Limit & Maturity Checks (TC-016 to TC-030) ---
    age_cases = [
        (18, 120, "HDFC Bank", "HOME_LOAN", "SALARIED", "REJECTED", "Age 18 is below the standard minimum entry age of 21."),
        (20, 120, "L&T Finance", "HOME_LOAN", "SALARIED", "REJECTED", "Age 20 is below L&T minimum entry age of 23."),
        (21, 180, "ICICI Bank", "HOME_LOAN", "SALARIED", "ELIGIBLE", "Age 21 is exactly on the ICICI Bank HL floor age."),
        (23, 240, "L&T Finance", "HOME_LOAN", "SALARIED", "ELIGIBLE", "Age 23 is exactly on the L&T HL floor age."),
        (25, 240, "Bank of Baroda", "HOME_LOAN", "SALARIED", "ELIGIBLE", "Age 25 is eligible. Matches standard age limits."),
        (35, 360, "SBI", "HOME_LOAN", "SALARIED", "ELIGIBLE", "Age 35 with 30 yr tenure. Maturity age is 65. Exactly on standard HL max limit."),
        (45, 240, "Yes Bank", "HOME_LOAN", "PROFESSIONAL", "ELIGIBLE", "Age 45 with 20 yr tenure. Maturity age is 65. Safe limit."),
        (55, 120, "HDFC Bank", "HOME_LOAN", "SALARIED", "ELIGIBLE", "Age 55 with 10 yr tenure. Maturity age is 65. Safe limit."),
        (58, 60, "HDFC Bank", "PERSONAL_LOAN", "SALARIED", "REJECTED", "Age 58 + 5 yr tenure = 63. Exceeds PL max age at maturity limit (60)."),
        (58, 240, "L&T Finance", "HOME_LOAN", "SALARIED", "REJECTED", "Age 58 + 20 yr tenure = 78. Exceeds L&T HL max age limit of 60 for salaried."),
        (59, 12, "L&T Finance", "HOME_LOAN", "SALARIED", "ELIGIBLE", "Age 59 + 1 yr tenure = 60. Exactly on the maximum age limit boundary for L&T salaried."),
        (61, 120, "ICICI Bank", "HOME_LOAN", "SELF_EMPLOYED", "ELIGIBLE", "Age 61 + 10 yr tenure = 71. Eligible since self-employed limit is 75 for ICICI."),
        (65, 120, "Bank of Baroda", "HOME_LOAN", "SELF_EMPLOYED", "ELIGIBLE", "Age 65 + 10 yr tenure = 75. Exactly on BOB self-employed limit of 75."),
        (71, 60, "Yes Bank", "HOME_LOAN", "SELF_EMPLOYED", "REJECTED", "Age 71 + 5 yr tenure = 76. Exceeds self-employed max maturity age (70)."),
        (75, 12, "SBI", "HOME_LOAN", "SALARIED", "REJECTED", "Age 75 exceeds all maximum age limits for home loans.")
    ]
    for idx, (age, tenure, lender, loan_type, emp, dec, reason) in enumerate(age_cases, 16):
        scenarios.append({
            "id": f"TC-{idx:03d}",
            "category": "Age & Maturity Limits",
            "objective": f"Verify maturity logic for age={age} and tenure={tenure} months with {lender}",
            "lender": lender,
            "product": loan_type,
            "employment": emp,
            "cibil": 750,
            "income": "Net: ₹100,000",
            "amount": 2000000,
            "tenure": tenure,
            "age": age,
            "expected_ltv": "75% to 80%" if dec == "ELIGIBLE" else "N/A",
            "expected_foir": "60% to 75%" if dec == "ELIGIBLE" else "N/A",
            "decision": dec,
            "reason": reason,
            "payload_template": {
                "lenderId": 105 if "HDFC" in lender else 73,
                "loanType": loan_type,
                "cibilScore": 750,
                "applicantAge": age,
                "employmentType": emp,
                "loanAmount": 2000000.00,
                "propertyValue": 3000000.00,
                "requestedTenureMonths": tenure,
                "monthlyIncome": 100000.00,
                "existingEmiTotal": 0.00,
                "idempotencyKey": f"test-age-{idx:03d}",
                "pinCode": "452001",
                "incomeComputationInput": {"programName": "NIP"}
            }
        })
        
    # --- Category 3: Income Floor Checks (TC-031 to TC-045) ---
    income_cases = [
        (10000, "Bandhan Bank", "HOME_LOAN", "SALARIED", "REJECTED", "Income ₹10,000 is below Bandhan's minimum floor of ₹15,000."),
        (14999, "Bandhan Bank", "HOME_LOAN", "SALARIED", "REJECTED", "Income ₹14,999 is below Bandhan's floor of ₹15,000."),
        (15000, "Bandhan Bank", "HOME_LOAN", "SALARIED", "ELIGIBLE", "Income ₹15,000 is exactly on Bandhan's minimum salary floor."),
        (20000, "ICICI Bank", "HOME_LOAN", "SALARIED", "REJECTED", "Income ₹20,000 is below ICICI NIP floor of ₹30,000."),
        (24999, "L&T Finance", "HOME_LOAN", "SELF_EMPLOYED", "REJECTED", "Income ₹24,999 is below L&T Self-Employed NIP floor of ₹25,000."),
        (25000, "L&T Finance", "HOME_LOAN", "SELF_EMPLOYED", "ELIGIBLE", "Income ₹25,000 matches L&T Self-Employed floor exactly."),
        (29999, "L&T Finance", "HOME_LOAN", "SALARIED", "REJECTED", "Income ₹29,999 is below L&T Salaried NIP floor of ₹30,000."),
        (30000, "L&T Finance", "HOME_LOAN", "SALARIED", "ELIGIBLE", "Income ₹30,000 matches L&T Salaried floor exactly."),
        (35000, "TATA Capital", "LOAN_AGAINST_PROPERTY", "SALARIED", "REJECTED", "Income ₹35,000 is below Tata Capital floor of ₹40,000."),
        (40000, "TATA Capital", "LOAN_AGAINST_PROPERTY", "SALARIED", "ELIGIBLE", "Income ₹40,000 matches Tata Capital LAP floor exactly."),
        (50000, "HDFC Bank", "HOME_LOAN", "SALARIED", "ELIGIBLE", "Income ₹50,000 is eligible. Standard NIP pricing applies."),
        (80000, "SBI", "HOME_LOAN", "SALARIED", "ELIGIBLE", "Income ₹80,000 is eligible. Standard NIP pricing applies."),
        (120000, "Yes Bank", "HOME_LOAN", "PROFESSIONAL", "ELIGIBLE", "Income ₹120,000 is eligible. SEP pricing applies."),
        (150000, "Aditya Birla Finance Limited", "LOAN_AGAINST_PROPERTY", "SELF_EMPLOYED", "ELIGIBLE", "Income ₹150,000 is eligible for High Leverage ABFL program."),
        (250000, "L&T Finance", "HOME_LOAN", "SALARIED", "ELIGIBLE", "Income ₹250,000 matches the highest FOIR bracket of L&T (80%).")
    ]
    for idx, (inc, lender, loan_type, emp, dec, reason) in enumerate(income_cases, 31):
        scenarios.append({
            "id": f"TC-{idx:03d}",
            "category": "Income Floor Validations",
            "objective": f"Verify minimum income floor checks for income={inc} with {lender}",
            "lender": lender,
            "product": loan_type,
            "employment": emp,
            "cibil": 750,
            "income": f"Net: ₹{inc:,}",
            "amount": 1000000,
            "tenure": 180,
            "age": 35,
            "expected_ltv": "75% to 80%" if dec == "ELIGIBLE" else "N/A",
            "expected_foir": "50% to 80%" if dec == "ELIGIBLE" else "N/A",
            "decision": dec,
            "reason": reason,
            "payload_template": {
                "lenderId": 73 if "L&T" in lender else 105,
                "loanType": loan_type,
                "cibilScore": 750,
                "applicantAge": 35,
                "employmentType": emp,
                "loanAmount": 1000000.00,
                "propertyValue": 1500000.00,
                "requestedTenureMonths": 180,
                "monthlyIncome": float(inc),
                "existingEmiTotal": 0.00,
                "idempotencyKey": f"test-income-{idx:03d}",
                "pinCode": "452001",
                "incomeComputationInput": {"programName": "NIP"}
            }
        })
        
    # --- Category 4: Employment & Program Routing (TC-046 to TC-060) ---
    routing_cases = [
        ("NIP", "SALARIED", "L&T Finance", "HOME_LOAN", "ELIGIBLE", "Routes to standard monthly Net Income Program. FOIR=60%."),
        ("SEP", "PROFESSIONAL", "L&T Finance", "HOME_LOAN", "ELIGIBLE", "Routes to CA multiplier (2.5x gross receipts). FOIR=75%."),
        ("SEP", "PROFESSIONAL", "Yes Bank", "HOME_LOAN", "ELIGIBLE", "Routes to Doctor program. FOIR=80%."),
        ("GST", "SELF_EMPLOYED", "L&T Finance", "HOME_LOAN", "ELIGIBLE", "GST Turnover program: 12% retail turnover is resolved as monthly income. FOIR=65%."),
        ("Banking", "SELF_EMPLOYED", "ICICI Bank", "LOAN_AGAINST_PROPERTY", "ELIGIBLE", "Banking ABB program: Uses Average Bank Balance to resolve monthly income. FOIR=33%."),
        ("CPM_SEP", "SELF_EMPLOYED", "Yes Bank", "LOAN_AGAINST_PROPERTY", "ELIGIBLE", "Cash Profit Method (PAT + Depreciation). FOIR=75%."),
        ("NIP", "SELF_EMPLOYED", "ICICI Bank", "LOAN_AGAINST_PROPERTY", "ELIGIBLE", "Dynamic NIP for self-employed: FOIR = 140 - LTV (1.40 - 0.60 = 0.80)."),
        ("SEP", "PROFESSIONAL", "Jio Finance", "HOME_LOAN", "ELIGIBLE", "CS Professional multiplier program (1.5x gross receipts). FOIR=70%."),
        ("GST", "SELF_EMPLOYED", "IDFC", "LOAN_AGAINST_PROPERTY", "ELIGIBLE", "IDFC GST Program with 75% FOIR matrix."),
        ("Banking", "SELF_EMPLOYED", "IDBI", "LOAN_AGAINST_PROPERTY", "ELIGIBLE", "IDBI Banking ABB Program with 60% FOIR matrix."),
        ("NIP", "SALARIED", "Bank of Baroda", "HOME_LOAN", "ELIGIBLE", "NIP Salaried routes to BOB standard matrix (60% to 75% FOIR)."),
        ("NIP", "SELF_EMPLOYED", "Bank of Baroda", "HOME_LOAN", "ELIGIBLE", "NIP Self Employed routes to BOB 70% FOIR lane, explicitly excluding professional 80% lane."),
        ("NIP", "PROFESSIONAL", "Bank of Baroda", "HOME_LOAN", "ELIGIBLE", "NIP Professional routes to BOB 80% FOIR lane, excluding self-employed 70% lane."),
        ("NIP", "SELF_EMPLOYED", "Aditya Birla Finance Limited", "LOAN_AGAINST_PROPERTY", "ELIGIBLE", "High leverage NIP Self-Employed, routes to 150% FOIR lane."),
        ("Low LTV", "SALARIED", "TATA Capital", "LOAN_AGAINST_PROPERTY", "ELIGIBLE", "Low LTV program: bypasses NIP/surrogate floor and caps LTV directly at 50%.")
    ]
    for idx, (prog, emp, lender, loan_type, dec, reason) in enumerate(routing_cases, 46):
        scenarios.append({
            "id": f"TC-{idx:03d}",
            "category": "Program & Route Checks",
            "objective": f"Verify program routing for program={prog} and employment={emp} with {lender}",
            "lender": lender,
            "product": loan_type,
            "employment": emp,
            "cibil": 740,
            "income": f"Program: {prog}",
            "amount": 3000000,
            "tenure": 180,
            "age": 35,
            "expected_ltv": "75%" if prog == "SEP" else "Variable",
            "expected_foir": "Variable",
            "decision": dec,
            "reason": reason,
            "payload_template": {
                "lenderId": 73 if "L&T" in lender else 200,
                "loanType": loan_type,
                "cibilScore": 740,
                "applicantAge": 35,
                "employmentType": emp,
                "loanAmount": 3000000.00,
                "propertyValue": 4500000.00,
                "requestedTenureMonths": 180,
                "monthlyIncome": 100000.00,
                "existingEmiTotal": 0.00,
                "idempotencyKey": f"test-routing-{idx:03d}",
                "pinCode": "452001",
                "incomeComputationInput": {
                    "programName": prog,
                    "grossReceipts": 3000000.00,
                    "profession": "CA" if prog == "SEP" else None,
                    "gstrTurnover12Months": 6000000.00 if prog == "GST" else None,
                    "averageBankBalance": 80000.00 if prog == "Banking" else None,
                    "pat": 1200000.00,
                    "depreciation": 100000.00
                }
            }
        })
        
    # --- Category 5: Property Type & Negative Lists (TC-061 to TC-075) ---
    property_cases = [
        ("FLAT", "Residential Category", "L&T Finance", "HOME_LOAN", "ELIGIBLE", "Standard Residential Flat is fully allowed. LTV resolved via HL grid."),
        ("HOME", "Residential Category", "ICICI Bank", "HOME_LOAN", "ELIGIBLE", "Residential Home is allowed. LTV resolved via HL grid."),
        ("VILLA", "Residential Category", "HDFC Bank", "HOME_LOAN", "ELIGIBLE", "Residential Villa is allowed. LTV resolved via HL grid."),
        ("SHOP", "Commercial Category", "TATA Capital", "LOAN_AGAINST_PROPERTY", "ELIGIBLE", "Commercial Shop allowed for TATA Capital LAP with LTV capped at 60%."),
        ("OFFICE", "Commercial Category", "ICICI Bank", "LOAN_AGAINST_PROPERTY", "ELIGIBLE", "Commercial Office allowed for ICICI LAP with LTV capped at 60%."),
        ("HOSPITAL", "Commercial Category", "Yes Bank", "LOAN_AGAINST_PROPERTY", "ELIGIBLE", "Commercial Hospital allowed for Yes Bank LAP with LTV capped at 60%."),
        ("PLOT", "Plot Category", "L&T Finance", "HOME_LOAN", "ELIGIBLE", "Plot loan is allowed. LTV is capped at flat 70% per HL Plot rule."),
        ("LAND", "Plot Category", "ICICI Bank", "HOME_LOAN", "ELIGIBLE", "Land loan is allowed. LTV is capped at flat 70% per HL Plot rule."),
        ("FACTORIES", "Industrial Category", "L&T Finance", "LOAN_AGAINST_PROPERTY", "ELIGIBLE", "Industrial Factory allowed for L&T LAP with LTV capped at 50%."),
        ("WAREHOUSE", "Industrial/Commercial Category", "TATA Capital", "LOAN_AGAINST_PROPERTY", "ELIGIBLE", "TATA Capital LAP allows Industrial Warehouses with LTV capped at 50%."),
        ("GODOWN", "Industrial/Commercial Category", "Bank of Baroda", "LOAN_AGAINST_PROPERTY", "ELIGIBLE", "BOB LAP allows Industrial Godown with LTV capped at 50%."),
        ("INFORMAL", "Negative Property List", "L&T Finance", "HOME_LOAN", "REJECTED", "Property sub-type matches negative property list (Informal is disallowed)."),
        ("Informal Property", "Negative Property List", "ICICI Bank", "HOME_LOAN", "REJECTED", "Property description resolves to negative property (Informal is disallowed)."),
        ("SCHOOL", "Negative Property List", "L&T Finance", "LOAN_AGAINST_PROPERTY", "REJECTED", "L&T Finance LAP explicitly denies School properties."),
        ("HOTEL", "Negative Property List", "Bajaj Prime", "LOAN_AGAINST_PROPERTY", "REJECTED", "Bajaj LAP explicitly denies Hotel properties.")
    ]
    for idx, (prop_type, cat, lender, loan_type, dec, reason) in enumerate(property_cases, 61):
        scenarios.append({
            "id": f"TC-{idx:03d}",
            "category": "Property Validations",
            "objective": f"Verify property type rules for propertyType='{prop_type}' with {lender}",
            "lender": lender,
            "product": loan_type,
            "employment": "SALARIED",
            "cibil": 740,
            "income": "Net: ₹100,000",
            "amount": 2500000,
            "tenure": 180,
            "age": 35,
            "expected_ltv": "Variable" if dec == "ELIGIBLE" else "N/A",
            "expected_foir": "65% to 80%" if dec == "ELIGIBLE" else "N/A",
            "decision": dec,
            "reason": reason,
            "payload_template": {
                "lenderId": 73 if "L&T" in lender else 105,
                "loanType": loan_type,
                "cibilScore": 740,
                "applicantAge": 35,
                "employmentType": "SALARIED",
                "propertyType": prop_type,
                "loanAmount": 2500000.00,
                "propertyValue": 4000000.00,
                "requestedTenureMonths": 180,
                "monthlyIncome": 100000.00,
                "existingEmiTotal": 0.00,
                "idempotencyKey": f"test-property-{idx:03d}",
                "pinCode": "452001",
                "incomeComputationInput": {"programName": "NIP"}
            }
        })
        
    # --- Category 6: Pin Code & Geo-Fencing (TC-076 to TC-090) ---
    pincodes_cases = [
        ("452001", "ELIGIBLE", "Indore City core. Fully inside service area."),
        ("452010", "ELIGIBLE", "Vijay Nagar Indore. Inside service area."),
        ("452020", "ELIGIBLE", "Bicholi Indore. Inside service area."),
        ("453001", "ELIGIBLE", "Indore district rural. Inside service area."),
        ("453441", "ELIGIBLE", "Mhow Indore district. Inside service area."),
        ("453551", "ELIGIBLE", "Sanwer Indore district. Inside service area."),
        ("452111", "ELIGIBLE", "Indore city peripheral. Inside service area."),
        ("453115", "ELIGIBLE", "Indore district rural. Inside service area."),
        ("462001", "REJECTED", "Bhopal PIN. Outside PRYME Indore service boundary."),
        ("400001", "REJECTED", "Mumbai PIN. Outside PRYME Indore service boundary."),
        ("110001", "REJECTED", "New Delhi PIN. Outside PRYME Indore service boundary."),
        ("560001", "REJECTED", "Bengaluru PIN. Outside PRYME Indore service boundary."),
        ("380001", "REJECTED", "Ahmedabad PIN. Outside PRYME Indore service boundary."),
        ("45200", "REJECTED", "Malformed PIN (5 digits). Must fail validation."),
        ("4520001", "REJECTED", "Malformed PIN (7 digits). Must fail validation.")
    ]
    for idx, (pin, dec, reason) in enumerate(pincodes_cases, 76):
        scenarios.append({
            "id": f"TC-{idx:03d}",
            "category": "Geo-Fence Checks",
            "objective": f"Verify geo-fence validation for pinCode='{pin}'",
            "lender": "L&T Finance",
            "product": "HOME_LOAN",
            "employment": "SALARIED",
            "cibil": 750,
            "income": "Net: ₹100,000",
            "amount": 2000000,
            "tenure": 240,
            "age": 35,
            "expected_ltv": "90% (LTV Grid)" if dec == "ELIGIBLE" else "N/A",
            "expected_foir": "75%" if dec == "ELIGIBLE" else "N/A",
            "decision": dec,
            "reason": reason,
            "payload_template": {
                "lenderId": 73,
                "loanType": "HOME_LOAN",
                "cibilScore": 750,
                "applicantAge": 35,
                "employmentType": "SALARIED",
                "loanAmount": 2000000.00,
                "propertyValue": 2500000.00,
                "requestedTenureMonths": 240,
                "monthlyIncome": 100000.00,
                "existingEmiTotal": 0.00,
                "idempotencyKey": f"test-pincode-{idx:03d}",
                "pinCode": pin,
                "incomeComputationInput": {"programName": "NIP"}
            }
        })
        
    # --- Category 7: SpEL & Decimal Boundary Check (TC-091 to TC-105) ---
    spel_boundary_cases = [
        ("60000.01", "ICICI Bank", "HOME_LOAN", "SALARIED", "ELIGIBLE", "60% FOIR (matches >= 60001.00 correctly due to .01 boundary check)."),
        ("60000.00", "ICICI Bank", "HOME_LOAN", "SALARIED", "ELIGIBLE", "50% FOIR (matches <= 60000.00 correctly. Decimals verified)."),
        ("30000.00", "ICICI Bank", "HOME_LOAN", "SALARIED", "ELIGIBLE", "50% FOIR (exactly on the lower salary boundary)."),
        ("29999.99", "ICICI Bank", "HOME_LOAN", "SALARIED", "REJECTED", "Below lowest salary floor of ₹30,000 (rejected on boundary check)."),
        ("100000.01", "ICICI Bank", "HOME_LOAN", "SALARIED", "ELIGIBLE", "65% FOIR (matches >= 100,001 slab accurately)."),
        ("100000.00", "ICICI Bank", "HOME_LOAN", "SALARIED", "ELIGIBLE", "60% FOIR (matches <= 100,000 slab accurately)."),
        ("150000.00", "L&T Finance", "HOME_LOAN", "SALARIED", "ELIGIBLE", "75% FOIR (matches <= 150000 bracket accurately)."),
        ("150000.01", "L&T Finance", "HOME_LOAN", "SALARIED", "ELIGIBLE", "80% FOIR (matches >= 150001 'No Limit' bracket accurately)."),
        ("250000.00", "L&T Finance", "HOME_LOAN", "SALARIED", "ELIGIBLE", "80% FOIR (matches 'No Limit' upper bound slab)."),
        ("3000000.00", "L&T Finance", "HOME_LOAN", "SALARIED", "ELIGIBLE", "90% LTV (requested loan amount exactly 30L is capped at 90% in L&T HL)."),
        ("3000000.50", "L&T Finance", "HOME_LOAN", "SALARIED", "ELIGIBLE", "80% LTV (requested loan amount slightly over 30L is capped at 80% in L&T HL)."),
        ("7500000.00", "L&T Finance", "HOME_LOAN", "SALARIED", "ELIGIBLE", "80% LTV (requested loan amount exactly 75L is capped at 80% in L&T HL)."),
        ("7500000.50", "L&T Finance", "HOME_LOAN", "SALARIED", "ELIGIBLE", "75% LTV (requested loan amount slightly over 75L is capped at 75% in L&T HL)."),
        ("Cash", "L&T Finance", "HOME_LOAN", "SALARIED", "REJECTED", "Salary mode check: 'Cash' is explicitly disallowed for L&T Home Loans."),
        ("UPI", "L&T Finance", "HOME_LOAN", "SALARIED", "REJECTED", "Salary mode check: 'UPI' is explicitly disallowed for L&T Home Loans.")
    ]
    for idx, (val, lender, loan_type, emp, dec, reason) in enumerate(spel_boundary_cases, 91):
        is_mode_case = val in ["Cash", "UPI"]
        monthly_inc = 100000.00 if is_mode_case else float(val)
        
        payload = {
            "lenderId": 73 if "L&T" in lender else 200,
            "loanType": loan_type,
            "cibilScore": 750,
            "applicantAge": 35,
            "employmentType": emp,
            "loanAmount": 3000000.50 if "3000000.50" in reason else (7500000.50 if "7500000.50" in reason else 2500000.00),
            "propertyValue": 4000000.00 if "3000000.50" in reason else (10000000.00 if "7500000.50" in reason else 3500000.00),
            "requestedTenureMonths": 240,
            "monthlyIncome": monthly_inc,
            "existingEmiTotal": 0.00,
            "idempotencyKey": f"test-spel-{idx:03d}",
            "pinCode": "452001",
            "incomeComputationInput": {"programName": "NIP"}
        }
        if is_mode_case:
            payload["negativeSalaryMode"] = val
            
        scenarios.append({
            "id": f"TC-{idx:03d}",
            "category": "SpEL & Boundary Edge Cases",
            "objective": f"Verify edge cases for input={val} with {lender}",
            "lender": lender,
            "product": loan_type,
            "employment": emp,
            "cibil": 750,
            "income": f"Mode: {val}" if is_mode_case else f"Net: ₹{monthly_inc:,}",
            "amount": payload["loanAmount"],
            "tenure": 240,
            "age": 35,
            "expected_ltv": "Variable" if dec == "ELIGIBLE" else "N/A",
            "expected_foir": "Variable" if dec == "ELIGIBLE" else "N/A",
            "decision": dec,
            "reason": reason,
            "payload_template": payload
        })
        
    # Verify count is 105
    print(f"Total scenarios generated: {len(scenarios)}")
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    # -------------------------------------------------------------
    # 3. WRITE TO TEST SCENARIOS SHEET
    # -------------------------------------------------------------
    ws_scenarios = wb.create_sheet(title="Test Scenarios")
    ws_scenarios.views.sheetView[0].showGridLines = True
    
    headers = [
        "Scenario ID", "Verify Objective", "Lender", "Product Type", "Employment",
        "CIBIL", "Income / Inputs", "Loan Amount", "Tenure (Mo)", "Age",
        "Expected LTV", "Expected FOIR", "Expected Decision", "Expected Logic & Reason"
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws_scenarios.cell(row=2, column=col_idx, value=header)
        cell.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, data in enumerate(scenarios, 3):
        ws_scenarios.cell(row=row_idx, column=1, value=data["id"]).alignment = Alignment(horizontal="center")
        ws_scenarios.cell(row=row_idx, column=2, value=data["objective"]).alignment = Alignment(wrap_text=True)
        ws_scenarios.cell(row=row_idx, column=3, value=data["lender"]).alignment = Alignment(horizontal="center")
        ws_scenarios.cell(row=row_idx, column=4, value=data["product"]).alignment = Alignment(horizontal="center")
        ws_scenarios.cell(row=row_idx, column=5, value=data["employment"]).alignment = Alignment(horizontal="center")
        ws_scenarios.cell(row=row_idx, column=6, value=data["cibil"]).alignment = Alignment(horizontal="center")
        ws_scenarios.cell(row=row_idx, column=7, value=data["income"]).alignment = Alignment(wrap_text=True)
        
        # Numeric / currency formatting
        amt_cell = ws_scenarios.cell(row=row_idx, column=8, value=data["amount"])
        amt_cell.number_format = '₹#,##0'
        amt_cell.alignment = Alignment(horizontal="right")
        
        ws_scenarios.cell(row=row_idx, column=9, value=data["tenure"]).alignment = Alignment(horizontal="center")
        ws_scenarios.cell(row=row_idx, column=10, value=data["age"]).alignment = Alignment(horizontal="center")
        
        ws_scenarios.cell(row=row_idx, column=11, value=data["expected_ltv"]).alignment = Alignment(horizontal="center")
        ws_scenarios.cell(row=row_idx, column=12, value=data["expected_foir"]).alignment = Alignment(horizontal="center")
        
        dec_cell = ws_scenarios.cell(row=row_idx, column=13, value=data["decision"])
        dec_cell.alignment = Alignment(horizontal="center")
        if data["decision"] == "ELIGIBLE":
            dec_cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            dec_cell.font = Font(name="Segoe UI", size=10, bold=True, color="375623")
        else:
            dec_cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            dec_cell.font = Font(name="Segoe UI", size=10, bold=True, color="C65911")
            
        ws_scenarios.cell(row=row_idx, column=14, value=data["reason"]).alignment = Alignment(wrap_text=True)
        
        for col in range(1, 15):
            c = ws_scenarios.cell(row=row_idx, column=col)
            c.font = Font(name="Segoe UI", size=10)
            c.border = thin_border
            
    # Auto-adjust column widths
    for col in ws_scenarios.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or '')
            lines = val.split('\n')
            for line in lines:
                if len(line) > max_len:
                    max_len = len(line)
        ws_scenarios.column_dimensions[col_letter].width = max(max_len + 3, 11)
        
    ws_scenarios.row_dimensions[2].height = 28
    
    # -------------------------------------------------------------
    # 4. WRITE TO JSON REQUEST PAYLOADS SHEET
    # -------------------------------------------------------------
    ws_payloads = wb.create_sheet(title="JSON Request Payloads")
    ws_payloads.views.sheetView[0].showGridLines = True
    
    ws_payloads.cell(row=2, column=1, value="Scenario ID").font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    ws_payloads.cell(row=2, column=1).fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    ws_payloads.cell(row=2, column=1).alignment = Alignment(horizontal="center")
    
    ws_payloads.cell(row=2, column=2, value="JSON Request Payload").font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    ws_payloads.cell(row=2, column=2).fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    
    for row_idx, data in enumerate(scenarios, 3):
        tc_label = f"{data['id']} ({data['lender']} - {data['product']})"
        id_cell = ws_payloads.cell(row=row_idx, column=1, value=tc_label)
        id_cell.font = Font(name="Segoe UI", size=10, bold=True)
        id_cell.alignment = Alignment(horizontal="center", vertical="top")
        id_cell.border = thin_border
        
        pretty_json = json.dumps(data["payload_template"], indent=2)
        payload_cell = ws_payloads.cell(row=row_idx, column=2, value=pretty_json)
        payload_cell.font = Font(name="Consolas", size=9)
        payload_cell.alignment = Alignment(vertical="top", wrap_text=True)
        payload_cell.border = thin_border
        
    ws_payloads.column_dimensions["A"].width = 32
    ws_payloads.column_dimensions["B"].width = 75
    
    # Save Workbook
    file_path = "/Users/manishmehta/Documents/PRYME-BACKEND-PROD/scratch/PRYME_Eligibility_Engine_Test_Scenarios.xlsx"
    wb.save(file_path)
    print(f"Excel file created successfully with 105 rows at: {file_path}")

if __name__ == "__main__":
    create_excel()
