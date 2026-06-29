#!/usr/bin/env python3
"""
═══════════════════════════════════════════════════════════════════════════════
PRYME Eligibility Engine — Excel ↔ Database Reconciliation Script
═══════════════════════════════════════════════════════════════════════════════
Reads 6 client Excel sheets (source of truth) and parses SQL migrations
(actual DB state), normalizes both into a standardized schema, and produces
a full discrepancy report.

Excel Sources (~/Downloads):
  1. eligibility workbook (1).xlsx  — Master eligibility conditions
  2. PF_data (1).xlsx               — Processing fee slabs
  3. FOIR_Sheet (1).xlsx            — FOIR slabs
  4. Login_fees (1).xlsx            — Login fee slabs
  5. HL_LTV_Sheet.xlsx              — Home Loan LTV grid
  6. LAP_LTV_Sheet.xlsx             — LAP LTV grid by property type

DB Source:
  SQL migrations V27–V35 in src/main/resources/db/migration/
═══════════════════════════════════════════════════════════════════════════════
"""

import os
import re
import sys
from pathlib import Path
from collections import defaultdict
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────

DOWNLOADS = Path.home() / "Downloads"
PROJECT   = Path.home() / "Documents" / "PRYME-BACKEND-PROD"
SQL_DIR   = PROJECT / "src" / "main" / "resources" / "db" / "migration"
OUTPUT    = PROJECT / "scratch" / "reconciliation_report.md"

EXCEL_FILES = {
    "eligibility": DOWNLOADS / "eligibility workbook (1).xlsx",
    "pf":          DOWNLOADS / "PF_data (1).xlsx",
    "foir":        DOWNLOADS / "FOIR_Sheet (1).xlsx",
    "login_fees":  DOWNLOADS / "Login_fees (1).xlsx",
    "hl_ltv":      DOWNLOADS / "HL_LTV_Sheet.xlsx",
    "lap_ltv":     DOWNLOADS / "LAP_LTV_Sheet.xlsx",
}

# Lender name normalization map (Excel → canonical)
LENDER_NORMALIZE = {
    "l&t finance":    "L&T Finance",
    "l&t":            "L&T Finance",
    "lt finance":     "L&T Finance",
    "icici bank":     "ICICI Bank",
    "icici":          "ICICI Bank",
    "bandhan bank":   "Bandhan Bank",
    "bandhan":        "Bandhan Bank",
    "aditya birla finance limited": "ABFL",
    "aditya birla":   "ABFL",
    "abfl":           "ABFL",
    "bank of baroda": "Bank of Baroda",
    "bob":            "Bank of Baroda",
    "sbi":            "SBI",
    "bajaj finance":  "Bajaj Finance",
    "bajaj prime":    "Bajaj Finance",
    "bajaj":          "Bajaj Finance",
    "bajaj affordable": "Bajaj Finance",
    "bajaj near prime": "Bajaj Finance",
    "yes bank":       "YES BANK",
    "yes":            "YES BANK",
    "hdfc bank":      "HDFC Bank",
    "hdfc":           "HDFC Bank",
    "jio finance":    "JIO Finance",
    "jio":            "JIO Finance",
    "idbi":           "IDBI",
    "tata capital":   "Tata Capital",
    "tata":           "Tata Capital",
    "idfc":           "IDFC",
    "idfc first bank": "IDFC",
    "icici hfc":      "ICICI HFC",
    "icicihfc":       "ICICI HFC",
}

def normalize_lender(name):
    """Normalize lender name to canonical form."""
    if not name:
        return None
    clean = str(name).strip()
    return LENDER_NORMALIZE.get(clean.lower(), clean)

def safe_str(val):
    """Convert value to string, handling None."""
    if val is None:
        return ""
    return str(val).strip()

def safe_num(val):
    """Convert value to number, handling 'No Limit', percentages, etc."""
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower() in ("na", "n/a", "-", ""):
        return None
    # Handle "No Limit" / "No limit" → infinity sentinel
    if s.lower().replace(" ", "") in ("nolimit", "nolimits", "unlimited"):
        return 999999999
    # Handle percentages
    if s.endswith("%"):
        try:
            return float(s.rstrip("%")) / 100.0
        except ValueError:
            return s
    # Handle comma-separated numbers
    s = s.replace(",", "").replace("₹", "").replace(" ", "")
    try:
        return float(s)
    except ValueError:
        return s


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL READERS
# ─────────────────────────────────────────────────────────────────────────────

def read_excel_sheet(filepath, sheet_name=None):
    """Read an Excel file into a list of dicts (one per row)."""
    wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [safe_str(h) for h in rows[0]]
    data = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        record = {}
        for i, h in enumerate(headers):
            if h and i < len(row):
                record[h] = row[i]
        data.append(record)
    return data

def read_all_sheets(filepath):
    """Read all sheets from an Excel file, returning {sheet_name: [rows]}."""
    wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
    result = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [safe_str(h) for h in rows[0]]
        data = []
        for row in rows[1:]:
            if all(v is None for v in row):
                continue
            record = {}
            for i, h in enumerate(headers):
                if h and i < len(row):
                    record[h] = row[i]
            data.append(record)
        result[name] = data
    wb.close()
    return result


# ─────────────────────────────────────────────────────────────────────────────
# SQL MIGRATION PARSER — Extract DB state from migration files
# ─────────────────────────────────────────────────────────────────────────────

def parse_product_inserts_v27():
    """Parse V27 which wipes and reseeds ALL loan_products."""
    filepath = SQL_DIR / "V27__ingest_master_roi_matrix.sql"
    if not filepath.exists():
        return {}
    text = filepath.read_text()
    products = {}
    # Match the VALUES block from the bulk INSERT
    pattern = re.compile(
        r"\('(\w[\w-]+)',\s*'([^']+)',\s*(\d+),\s*'([^']+)',\s*'(\w+)',\s*'(\w+)',\s*"
        r"(\d+),\s*(\d+),\s*([\d.]+),\s*(\d+),\s*(\d+),\s*(\d+),\s*(\d+),\s*"
        r"(true|false),\s*([\d.]+),\s*(\d+)\)"
    )
    for m in pattern.finditer(text):
        code = m.group(1)
        products[code] = {
            "product_code": code,
            "product_name": m.group(2),
            "lender_id": int(m.group(3)),
            "lender_name": m.group(4),
            "loan_type": m.group(5),
            "interest_type": m.group(6),
            "min_cibil": int(m.group(7)),
            "max_cibil": int(m.group(8)),
            "roi": float(m.group(9)),
            "min_loan_amount": int(m.group(10)),
            "max_loan_amount": int(m.group(11)),
            "min_tenure_months": int(m.group(12)),
            "max_tenure_months": int(m.group(13)),
            "is_active": m.group(14) == "true",
            "processing_fee": float(m.group(15)),
            "login_fees": int(m.group(16)),
        }
    return products

def parse_product_inserts_v29():
    """Parse V29 which adds JIO, IDBI, TATA, IDFC, ICICI HFC products."""
    filepath = SQL_DIR / "V29__create_and_seed_pf_matrix.sql"
    if not filepath.exists():
        return {}
    text = filepath.read_text()
    products = {}
    pattern = re.compile(
        r"\('(\w[\w-]+)',\s*'([^']+)',\s*(\d+),\s*'([^']+)',\s*'(\w+)',\s*'(\w+)',\s*"
        r"(\d+),\s*(\d+),\s*([\d.]+),\s*(\d+),\s*(\d+),\s*(\d+),\s*(\d+),\s*"
        r"(true|false),\s*([\d.]+),\s*(\d+)\)"
    )
    for m in pattern.finditer(text):
        code = m.group(1)
        products[code] = {
            "product_code": code,
            "product_name": m.group(2),
            "lender_id": int(m.group(3)),
            "lender_name": m.group(4),
            "loan_type": m.group(5),
            "interest_type": m.group(6),
            "min_cibil": int(m.group(7)),
            "max_cibil": int(m.group(8)),
            "roi": float(m.group(9)),
            "min_loan_amount": int(m.group(10)),
            "max_loan_amount": int(m.group(11)),
            "min_tenure_months": int(m.group(12)),
            "max_tenure_months": int(m.group(13)),
            "is_active": m.group(14) == "true",
            "processing_fee": float(m.group(15)),
            "login_fees": int(m.group(16)),
        }
    return products

def parse_foir_inserts_v28():
    """Parse V28 FOIR slabs ingestion (uses existing eligibility_conditions schema)."""
    filepath = SQL_DIR / "V28__ingest_foir_slabs_existing_schema.sql"
    if not filepath.exists():
        return []
    text = filepath.read_text()
    foir_rows = []
    # Match INSERT patterns for FOIR data
    # These are inserted as eligibility_conditions with foir_max set
    pattern = re.compile(
        r"product_code\s*=\s*'([\w-]+)'.*?employment_type\s*=\s*'([^']*)'.*?"
        r"surrogate\s*=\s*'([^']*)'.*?foir_max\s*=\s*([\d.]+)",
        re.DOTALL
    )
    for m in pattern.finditer(text):
        foir_rows.append({
            "product_code": m.group(1),
            "employment_type": m.group(2),
            "surrogate": m.group(3),
            "foir_max": float(m.group(4)),
        })
    return foir_rows

def parse_enrichment_v32():
    """Parse V32 eligibility conditions enrichment updates."""
    filepath = SQL_DIR / "V32__enrich_product_config_master.sql"
    if not filepath.exists():
        return {}
    text = filepath.read_text()
    enrichments = {}
    # Extract fee UPDATEs for loan_products
    fee_pattern = re.compile(
        r"UPDATE loan_products SET\s+(.*?)\s+WHERE\s+product_code\s+LIKE\s+'([^']+)'",
        re.DOTALL
    )
    for m in fee_pattern.finditer(text):
        sets_str = m.group(1)
        code_pattern = m.group(2)  # e.g. 'LT-HL%'
        fees = {}
        for kv in sets_str.split(","):
            kv = kv.strip()
            if "=" in kv:
                k, v = kv.split("=", 1)
                k = k.strip()
                v = v.strip().strip("'")
                if v == "NULL":
                    v = None
                fees[k] = v
        enrichments[code_pattern] = fees
    return enrichments


# ─────────────────────────────────────────────────────────────────────────────
# NORMALIZATION — Convert Excel "wide" → standardized "tall" format
# ─────────────────────────────────────────────────────────────────────────────

def normalize_eligibility_excel(rows):
    """
    Normalize eligibility workbook rows into:
    (lender, loan_type, emp_type, surrogate, attribute, value)
    """
    records = []
    for row in rows:
        lender = normalize_lender(row.get("Lender_Name"))
        loan_type = safe_str(row.get("Product_Name") or row.get("Loan_Type", ""))
        # Normalize loan_type
        lt = loan_type.upper().strip()
        if lt in ("HL", "HOME LOAN"):
            loan_type = "HL"
        elif lt in ("LAP", "LOAN AGAINST PROPERTY"):
            loan_type = "LAP"

        emp_type = safe_str(row.get("Employment_Type", ""))
        surrogate = safe_str(row.get("Surrogate", "NIP"))
        sep_prof = safe_str(row.get("Self Employed Professional", ""))

        key = (lender, loan_type, emp_type, surrogate, sep_prof)

        # Extract all attribute-value pairs
        skip_keys = {"Product_Name", "Loan_Type", "Lender_Name", "Employment_Type",
                     "Surrogate", "Self Employed Professional"}
        for attr, val in row.items():
            if attr in skip_keys or val is None:
                continue
            sval = safe_str(val)
            if sval and sval.lower() not in ("", "na", "n/a"):
                records.append({
                    "lender": lender,
                    "loan_type": loan_type,
                    "emp_type": emp_type,
                    "surrogate": surrogate,
                    "sep_profession": sep_prof,
                    "attribute": attr,
                    "excel_value": sval,
                })
    return records

def normalize_foir_excel(rows):
    """Normalize FOIR sheet rows."""
    records = []
    for row in rows:
        lender = normalize_lender(row.get("Lender_Name"))
        loan_type = safe_str(row.get("Product_Name", "")).upper().strip()
        if loan_type in ("HL", "HOME LOAN"):
            loan_type = "HL"
        elif loan_type in ("LAP", "LOAN AGAINST PROPERTY"):
            loan_type = "LAP"

        emp_type = safe_str(row.get("Employement_Type") or row.get("Employment_Type", ""))
        surrogate = safe_str(row.get("Surrogate", "NIP"))
        lower_sal = row.get("Lower_Salary")
        upper_sal = row.get("Upper_Salary")
        foir = row.get("FOIR (%)")
        deviation = row.get("Deviation")

        records.append({
            "lender": lender,
            "loan_type": loan_type,
            "emp_type": emp_type,
            "surrogate": surrogate,
            "lower_salary": safe_num(lower_sal),
            "upper_salary": safe_num(upper_sal),
            "foir_pct": safe_str(foir),
            "deviation": safe_str(deviation),
        })
    return records

def normalize_pf_excel(rows):
    """Normalize PF_data sheet rows."""
    records = []
    for row in rows:
        lender = normalize_lender(row.get("Lender_Name"))
        loan_type = safe_str(row.get("Product_Name") or row.get("Loan_Type", "")).upper().strip()
        if loan_type in ("HL", "HOME LOAN"):
            loan_type = "HL"
        elif loan_type in ("LAP", "LOAN AGAINST PROPERTY"):
            loan_type = "LAP"

        records.append({
            "lender": lender,
            "loan_type": loan_type,
            "emp_type": safe_str(row.get("Employment_Type", "")),
            "min_loan": safe_num(row.get("Min_Loan_Amount")),
            "max_loan": safe_num(row.get("Max_Loan_Amount")),
            "pf": safe_str(row.get("PF", "")),
            "tax": safe_str(row.get("Tax", "")),
            "notes": safe_str(row.get("Notes", "")),
        })
    return records

def normalize_login_fees_excel(rows):
    """Normalize Login_fees sheet rows."""
    records = []
    for row in rows:
        lender = normalize_lender(row.get("Lender_Name"))
        loan_type = safe_str(row.get("Product_Name") or row.get("Loan_Type", "")).upper().strip()
        if loan_type in ("HL", "HOME LOAN"):
            loan_type = "HL"
        elif loan_type in ("LAP", "LOAN AGAINST PROPERTY"):
            loan_type = "LAP"

        records.append({
            "lender": lender,
            "loan_type": loan_type,
            "emp_type": safe_str(row.get("Employment_Type", "")),
            "min_loan": safe_num(row.get("Min_Loan_Amount")),
            "max_loan": safe_num(row.get("Max_Loan_Amount")),
            "login_fee": safe_str(row.get("Login Fees", "")),
        })
    return records

def normalize_ltv_excel(rows, loan_type_label):
    """Normalize LTV sheet rows (both HL and LAP)."""
    records = []
    for row in rows:
        lender = normalize_lender(row.get("Lender_Name"))
        for prop_key, val in row.items():
            if prop_key == "Lender_Name" or val is None:
                continue
            sval = safe_str(val)
            if sval.lower() == "negative":
                ltv = "NEGATIVE"
            else:
                ltv = sval
            records.append({
                "lender": lender,
                "loan_type": loan_type_label,
                "property_type": prop_key.strip(),
                "ltv_value": ltv,
            })
    return records


# ─────────────────────────────────────────────────────────────────────────────
# BUILD DB STATE — Parse SQL migrations to build "actual" state
# ─────────────────────────────────────────────────────────────────────────────

def parse_and_apply_v31_login_fees(products):
    """Parse V31 login fee matrix table and update products login_fees field."""
    filepath = SQL_DIR / "V31__create_and_seed_login_fee_matrix.sql"
    if not filepath.exists():
        return
    text = filepath.read_text()
    
    # Split text by the SELECT statement to isolate each product's inserts
    blocks = re.split(r"SELECT id INTO p_id FROM loan_products WHERE product_code =", text, flags=re.IGNORECASE)
    for block in blocks[1:]:
        m_code = re.match(r"\s*'([^']+)'", block)
        if not m_code:
            continue
        code = m_code.group(1)
        
        # Match the INSERT statement and extract the last value (login_fee)
        m_insert = re.search(r"INSERT INTO.*?\((.*?)\);", block, re.DOTALL | re.IGNORECASE)
        if m_insert and code in products:
            values_part = m_insert.group(1).strip()
            last_val = values_part.split(",")[-1].strip()
            try:
                products[code]["login_fees"] = float(last_val)
            except ValueError:
                pass

def parse_and_apply_v36_updates(products):
    """Parse V36 database updates for loan_products and apply to the catalog."""
    filepath = SQL_DIR / "V36__align_db_with_client_excel.sql"
    if not filepath.exists():
        return
    text = filepath.read_text()
    pattern = re.compile(r"UPDATE loan_products SET\s+(.*?)\s+WHERE\s+product_code\s+=\s+'([^']+)';", re.IGNORECASE)
    for m in pattern.finditer(text):
        set_clause = m.group(1)
        code = m.group(2)
        if code in products:
            parts = [p.strip() for p in set_clause.split(",")]
            for part in parts:
                if "=" in part:
                    k, v = part.split("=", 1)
                    k = k.strip().lower()
                    v = v.strip()
                    if v.lower() == "null":
                        products[code][k] = None
                    else:
                        try:
                            if "." in v:
                                products[code][k] = float(v)
                            else:
                                products[code][k] = int(v)
                        except ValueError:
                            products[code][k] = v.strip("'")

def build_db_products():
    """Build complete product catalog from V27 + V29, updated by V31 (login fees) and V36 (alignment)."""
    products = parse_product_inserts_v27()
    products.update(parse_product_inserts_v29())
    parse_and_apply_v31_login_fees(products)
    parse_and_apply_v36_updates(products)
    return products

def get_db_lenders(products):
    """Get unique lenders from DB products."""
    lenders = set()
    for p in products.values():
        lenders.add(p["lender_name"])
    return lenders

def get_db_loan_types(products):
    """Get loan types per lender from DB."""
    result = defaultdict(set)
    for p in products.values():
        result[p["lender_name"]].add(p["loan_type"])
    return result


# ─────────────────────────────────────────────────────────────────────────────
# COMPARISON ENGINE
# ─────────────────────────────────────────────────────────────────────────────

def compare_products(excel_elig, db_products):
    """
    Compare products between Excel eligibility workbook and DB.
    Returns: (missing_from_db, extra_in_db, common)
    """
    # Build Excel product set: (lender, loan_type) pairs
    excel_products = set()
    for row in excel_elig:
        lender = normalize_lender(row.get("Lender_Name"))
        loan_type = safe_str(row.get("Product_Name", "")).upper().strip()
        if loan_type in ("HL", "HOME LOAN"):
            loan_type = "HL"
        elif loan_type in ("LAP", "LOAN AGAINST PROPERTY"):
            loan_type = "LAP"
        if lender and loan_type:
            excel_products.add((lender, loan_type))

    # Build DB product set: (normalized_lender, loan_type) pairs
    db_product_set = set()
    for p in db_products.values():
        norm_lender = normalize_lender(p["lender_name"])
        db_product_set.add((norm_lender, p["loan_type"]))

    missing_from_db = excel_products - db_product_set
    extra_in_db = db_product_set - excel_products
    common = excel_products & db_product_set

    return missing_from_db, extra_in_db, common

def compare_eligibility_attributes(excel_elig, db_products):
    """
    Compare detailed eligibility attributes (CIBIL, tenure, amounts, ages, etc.)
    between Excel and DB for each lender × loan_type.
    """
    discrepancies = []

    # Group Excel rows by (lender, loan_type, emp_type, surrogate)
    excel_by_key = defaultdict(list)
    for row in excel_elig:
        lender = normalize_lender(row.get("Lender_Name"))
        lt = safe_str(row.get("Product_Name", "")).upper().strip()
        if lt in ("HL", "HOME LOAN"):
            lt = "HL"
        elif lt in ("LAP", "LOAN AGAINST PROPERTY"):
            lt = "LAP"
        emp = safe_str(row.get("Employment_Type", ""))
        surr = safe_str(row.get("Surrogate", "NIP"))
        excel_by_key[(lender, lt, emp, surr)].append(row)

    # Compare key numeric fields against DB products
    attr_map = {
        "MIN_CIBIL": "min_cibil",
        "Min_Tenure (Months)": "min_tenure_months",
        "Max_Tenure (Months)": "max_tenure_months",
        "Min_LoanAmount": "min_loan_amount",
        "Max_LoanAmount": "max_loan_amount",
        "Min_Age": None,  # in eligibility_conditions, not products
        "Max_Age": None,
        "Min_Income": None,
    }

    for (lender, lt, emp, surr), rows in excel_by_key.items():
        if surr != "NIP":
            continue  # Only compare base NIP product params

        # Find matching DB product
        matching_db = []
        for p in db_products.values():
            if normalize_lender(p["lender_name"]) == lender and p["loan_type"] == lt:
                matching_db.append(p)

        if not matching_db:
            continue

        # Filter matching_db by employment type compatibility to prevent false matches
        db_emp_type = "salaried" if "salaried" in emp.lower() and "self" not in emp.lower() else "self employed"
        selected_p = None
        for p in matching_db:
            p_emp = "salaried" if "-0001" in p["product_code"] else "self employed"
            if p_emp == db_emp_type:
                selected_p = p
                break
        if not selected_p:
            selected_p = matching_db[0]

        # Aggregate Excel values for this group to get the base product-level limit range
        # (Union of all property type rows)
        group_excel_limits = {}
        for excel_attr, db_attr in attr_map.items():
            if not db_attr:
                continue
            vals = []
            for r in rows:
                val = r.get(excel_attr)
                if val is not None:
                    num = safe_num(val)
                    if num is not None and not isinstance(num, str):
                        vals.append(num)
            if vals:
                if excel_attr in ("MIN_CIBIL", "Min_Tenure (Months)", "Min_LoanAmount"):
                    group_excel_limits[excel_attr] = min(vals)
                else:
                    group_excel_limits[excel_attr] = max(vals)

        for excel_attr, db_attr in attr_map.items():
            if not db_attr:
                continue
            excel_num = group_excel_limits.get(excel_attr)
            db_val = selected_p.get(db_attr)
            if excel_num is not None and db_val is not None:
                # Compare
                try:
                    e = float(excel_num)
                    d = float(db_val)
                    if abs(e - d) > 0.001:
                        discrepancies.append({
                            "lender": lender,
                            "loan_type": lt,
                            "emp_type": emp,
                            "surrogate": surr,
                            "attribute": excel_attr,
                            "excel_value": str(excel_num),
                            "db_value": str(db_val),
                            "severity": "HIGH" if excel_attr in ("MIN_CIBIL", "Min_LoanAmount", "Max_LoanAmount") else "MEDIUM",
                        })
                except (ValueError, TypeError):
                    pass

    return discrepancies

def compare_foir(excel_foir_rows, db_foir_file):
    """Compare FOIR slabs between Excel and V28 SQL migration."""
    discrepancies = []
    
    if not db_foir_file.exists():
        discrepancies.append({
            "lender": "ALL",
            "loan_type": "ALL",
            "attribute": "FOIR_MIGRATION",
            "excel_value": f"{len(excel_foir_rows)} rows",
            "db_value": "V28 migration file not found",
            "severity": "CRITICAL",
        })
        return discrepancies

    db_text = db_foir_file.read_text()

    # For each Excel FOIR row, check if there's a corresponding entry in V28
    for row in excel_foir_rows:
        lender = row["lender"]
        lt = row["loan_type"]
        foir = row["foir_pct"]
        surr = row["surrogate"]

        if not lender or not foir:
            continue

        # Normalize FOIR value for search
        foir_str = safe_str(foir)
        if foir_str and "%" in foir_str:
            try:
                foir_decimal = float(foir_str.rstrip("%")) / 100.0
            except ValueError:
                foir_decimal = None
        elif foir_str:
            try:
                foir_decimal = float(foir_str)
                if foir_decimal > 1:
                    foir_decimal = foir_decimal / 100.0
            except ValueError:
                foir_decimal = None
        else:
            foir_decimal = None

        # Check if this lender appears in V28 at all
        lender_patterns = [lender]
        if lender == "L&T Finance":
            lender_patterns.extend(["LT-", "L&T"])
        elif lender == "ICICI Bank":
            lender_patterns.extend(["ICICI-"])
        elif lender == "Bajaj Finance":
            lender_patterns.extend(["BAJAJ-"])
        elif lender == "YES BANK":
            lender_patterns.extend(["YES-"])
        elif lender == "HDFC Bank":
            lender_patterns.extend(["HDFC-"])
        elif lender == "JIO Finance":
            lender_patterns.extend(["JIO-"])
        elif lender == "Tata Capital":
            lender_patterns.extend(["TATA-"])
        elif lender == "ABFL":
            lender_patterns.extend(["ABFL-"])
        elif lender == "Bank of Baroda":
            lender_patterns.extend(["BOB-"])
        elif lender == "Bandhan Bank":
            lender_patterns.extend(["BANDHAN-"])
        elif lender == "IDBI":
            lender_patterns.extend(["IDBI-"])
        elif lender == "IDFC":
            lender_patterns.extend(["IDFC-"])

        found = False
        for pat in lender_patterns:
            if pat in db_text and lt in db_text:
                found = True
                break

        if not found:
            discrepancies.append({
                "lender": lender,
                "loan_type": lt,
                "attribute": f"FOIR_{surr}",
                "excel_value": foir_str,
                "db_value": "NOT FOUND in V28",
                "severity": "HIGH",
            })

    return discrepancies

def compare_login_fees(excel_login, db_products):
    """Compare login fees between Excel and DB products."""
    discrepancies = []
    for row in excel_login:
        lender = row["lender"]
        lt = row["loan_type"]
        login_fee_str = row["login_fee"]

        if not lender or not login_fee_str:
            continue

        try:
            excel_fee = float(safe_num(login_fee_str))
        except (ValueError, TypeError):
            continue

        # Find matching DB product(s)
        matching = [
            p for p in db_products.values()
            if normalize_lender(p["lender_name"]) == lender and p["loan_type"] == lt
        ]

        if not matching:
            discrepancies.append({
                "lender": lender,
                "loan_type": lt,
                "attribute": "Login_Fee",
                "excel_value": login_fee_str,
                "db_value": "NO PRODUCT FOUND",
                "severity": "HIGH",
            })
            continue

        for mp in matching:
            # Check employment type compatibility to prevent false matches between salaried/self-employed variants
            db_emp = "salaried" if "-0001" in mp["product_code"] else "self employed"
            excel_emp = row["emp_type"].lower()
            if ("salaried" in excel_emp and db_emp != "salaried") or ("self" in excel_emp and db_emp != "self employed"):
                continue

            db_fee = mp.get("login_fees", 0)
            if db_fee is None:
                db_fee = 0
            if abs(excel_fee - db_fee) > 0.5:
                discrepancies.append({
                    "lender": lender,
                    "loan_type": lt,
                    "attribute": f"Login_Fee ({mp['product_code']})",
                    "excel_value": login_fee_str,
                    "db_value": str(db_fee),
                    "severity": "MEDIUM",
                })

    return discrepancies

def compare_pf(excel_pf, db_pf_file):
    """Compare processing fees between Excel and V29 PF matrix."""
    discrepancies = []
    if not db_pf_file.exists():
        discrepancies.append({
            "lender": "ALL",
            "loan_type": "ALL",
            "attribute": "PF_MIGRATION",
            "excel_value": f"{len(excel_pf)} rows",
            "db_value": "V29 migration file not found",
            "severity": "CRITICAL",
        })
        return discrepancies

    db_text = db_pf_file.read_text()
    
    for row in excel_pf:
        lender = row["lender"]
        lt = row["loan_type"]
        pf = row["pf"]

        if not lender or not pf:
            continue

        # Check lender presence in PF matrix
        lender_code_map = {
            "L&T Finance": "LT-",
            "ICICI Bank": "ICICI-",
            "Bajaj Finance": "BAJAJ-",
            "YES BANK": "YES-",
            "HDFC Bank": "HDFC-",
            "JIO Finance": "JIO-",
            "Tata Capital": "TATA-",
            "ABFL": "ABFL-",
            "Bank of Baroda": "BOB-",
            "Bandhan Bank": "BANDHAN-",
            "IDBI": "IDBI-",
            "IDFC": "IDFC-",
            "SBI": "SBI-",
            "ICICI HFC": "ICICIHFC-",
        }
        code_prefix = lender_code_map.get(lender, "")
        search_code = f"{code_prefix}{lt}"

        if code_prefix and search_code not in db_text and code_prefix not in db_text:
            discrepancies.append({
                "lender": lender,
                "loan_type": lt,
                "attribute": "PF_Matrix",
                "excel_value": pf,
                "db_value": f"NOT FOUND (searched for {search_code})",
                "severity": "HIGH",
            })

    return discrepancies


# ─────────────────────────────────────────────────────────────────────────────
# REPORT GENERATOR
# ─────────────────────────────────────────────────────────────────────────────

def generate_report(
    excel_products_set, db_products_set, missing, extra, common,
    elig_discrepancies, foir_discrepancies, login_fee_discrepancies,
    pf_discrepancies, excel_elig_rows, excel_foir_rows, excel_pf_rows,
    excel_login_rows, excel_hl_ltv_rows, excel_lap_ltv_rows,
    db_products
):
    """Generate comprehensive markdown report."""
    lines = []
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    lines.append(f"# PRYME Engine — Excel ↔ Database Reconciliation Report")
    lines.append(f"\n> Generated: {now}")
    lines.append(f"> Excel Sheets: 6 files from ~/Downloads")
    lines.append(f"> DB Source: SQL migrations V27–V35\n")
    lines.append("---\n")

    # ── Summary
    lines.append("## 1. Executive Summary\n")
    total_disc = len(elig_discrepancies) + len(foir_discrepancies) + len(login_fee_discrepancies) + len(pf_discrepancies)
    lines.append(f"| Metric | Count |")
    lines.append(f"|--------|-------|")
    lines.append(f"| Excel Lender×LoanType pairs | {len(excel_products_set)} |")
    lines.append(f"| DB Lender×LoanType pairs | {len(db_products_set)} |")
    lines.append(f"| **Missing from DB** | **{len(missing)}** |")
    lines.append(f"| **Extra in DB** | **{len(extra)}** |")
    lines.append(f"| Common (matched) | {len(common)} |")
    lines.append(f"| **Total Discrepancies** | **{total_disc}** |")
    lines.append(f"| Eligibility Workbook rows | {len(excel_elig_rows)} |")
    lines.append(f"| FOIR Sheet rows | {len(excel_foir_rows)} |")
    lines.append(f"| PF Sheet rows | {len(excel_pf_rows)} |")
    lines.append(f"| Login Fee Sheet rows | {len(excel_login_rows)} |")
    lines.append(f"| HL LTV Sheet rows | {len(excel_hl_ltv_rows)} |")
    lines.append(f"| LAP LTV Sheet rows | {len(excel_lap_ltv_rows)} |")
    lines.append(f"| DB Product Codes | {len(db_products)} |")
    lines.append("")

    # ── Section 2: Missing Products
    lines.append("---\n")
    lines.append("## 2. Products in Excel but MISSING from Database\n")
    if missing:
        lines.append("> [!CAUTION]")
        lines.append("> These lender×loan-type combinations exist in the client's Excel sheets but have NO corresponding product in the database.\n")
        lines.append("| # | Lender | Loan Type | Status |")
        lines.append("|---|--------|-----------|--------|")
        for i, (lender, lt) in enumerate(sorted(missing), 1):
            lines.append(f"| {i} | {lender} | {lt} | 🔴 MISSING |")
    else:
        lines.append("✅ No products missing from DB — all Excel products have DB entries.\n")
    lines.append("")

    # ── Section 3: Extra Products
    lines.append("---\n")
    lines.append("## 3. Products in Database but NOT in Excel (Extra/Stale)\n")
    if extra:
        lines.append("> [!WARNING]")
        lines.append("> These lender×loan-type combinations exist in the DB but are NOT in the client's Excel sheets. They may be stale or test data.\n")
        lines.append("| # | Lender | Loan Type | Status |")
        lines.append("|---|--------|-----------|--------|")
        for i, (lender, lt) in enumerate(sorted(extra), 1):
            lines.append(f"| {i} | {lender} | {lt} | ⚠️ EXTRA |")
    else:
        lines.append("✅ No extra products in DB — database matches Excel exactly.\n")
    lines.append("")

    # ── Section 4: Eligibility Attribute Discrepancies
    lines.append("---\n")
    lines.append("## 4. Eligibility Attribute Discrepancies (Excel vs DB)\n")
    if elig_discrepancies:
        lines.append("> [!IMPORTANT]")
        lines.append("> These fields have different values in Excel vs Database.\n")
        lines.append("| # | Lender | Loan Type | Emp Type | Attribute | Excel Value | DB Value | Severity |")
        lines.append("|---|--------|-----------|----------|-----------|-------------|----------|----------|")
        for i, d in enumerate(sorted(elig_discrepancies, key=lambda x: x["severity"]), 1):
            sev_icon = "🔴" if d["severity"] == "HIGH" else "🟡" if d["severity"] == "MEDIUM" else "🟢"
            lines.append(f"| {i} | {d['lender']} | {d['loan_type']} | {d['emp_type']} | {d['attribute']} | {d['excel_value']} | {d['db_value']} | {sev_icon} {d['severity']} |")
    else:
        lines.append("✅ All eligibility attributes match between Excel and DB.\n")
    lines.append("")

    # ── Section 5: FOIR Discrepancies
    lines.append("---\n")
    lines.append("## 5. FOIR Slab Discrepancies\n")
    if foir_discrepancies:
        lines.append("| # | Lender | Loan Type | Attribute | Excel Value | DB Value | Severity |")
        lines.append("|---|--------|-----------|-----------|-------------|----------|----------|")
        for i, d in enumerate(sorted(foir_discrepancies, key=lambda x: x["severity"]), 1):
            sev_icon = "🔴" if d["severity"] == "HIGH" else "🟡"
            lines.append(f"| {i} | {d['lender']} | {d['loan_type']} | {d['attribute']} | {d['excel_value']} | {d['db_value']} | {sev_icon} {d['severity']} |")
    else:
        lines.append("✅ All FOIR slabs have corresponding DB entries.\n")
    lines.append("")

    # ── Section 6: Login Fee Discrepancies
    lines.append("---\n")
    lines.append("## 6. Login Fee Discrepancies\n")
    if login_fee_discrepancies:
        lines.append("| # | Lender | Loan Type | Attribute | Excel Value | DB Value | Severity |")
        lines.append("|---|--------|-----------|-----------|-------------|----------|----------|")
        for i, d in enumerate(sorted(login_fee_discrepancies, key=lambda x: x["severity"]), 1):
            sev_icon = "🔴" if d["severity"] == "HIGH" else "🟡"
            lines.append(f"| {i} | {d['lender']} | {d['loan_type']} | {d['attribute']} | {d['excel_value']} | {d['db_value']} | {sev_icon} {d['severity']} |")
    else:
        lines.append("✅ All login fees match.\n")
    lines.append("")

    # ── Section 7: Processing Fee Discrepancies
    lines.append("---\n")
    lines.append("## 7. Processing Fee (PF) Discrepancies\n")
    if pf_discrepancies:
        lines.append("| # | Lender | Loan Type | Attribute | Excel Value | DB Value | Severity |")
        lines.append("|---|--------|-----------|-----------|-------------|----------|----------|")
        for i, d in enumerate(sorted(pf_discrepancies, key=lambda x: x["severity"]), 1):
            sev_icon = "🔴" if d["severity"] == "HIGH" else "🟡"
            lines.append(f"| {i} | {d['lender']} | {d['loan_type']} | {d['attribute']} | {d['excel_value']} | {d['db_value']} | {sev_icon} {d['severity']} |")
    else:
        lines.append("✅ All PF entries have corresponding DB entries.\n")
    lines.append("")

    # ── Section 8: Full Product Catalog Cross-Reference
    lines.append("---\n")
    lines.append("## 8. Full Product Catalog Cross-Reference\n")
    all_pairs = sorted(excel_products_set | db_products_set)
    lines.append("| Lender | Loan Type | In Excel | In DB | Status |")
    lines.append("|--------|-----------|----------|-------|--------|")
    for lender, lt in all_pairs:
        in_excel = "✅" if (lender, lt) in excel_products_set else "❌"
        in_db = "✅" if (lender, lt) in db_products_set else "❌"
        if (lender, lt) in missing:
            status = "🔴 MISSING FROM DB"
        elif (lender, lt) in extra:
            status = "⚠️ EXTRA IN DB"
        else:
            status = "✅ MATCHED"
        lines.append(f"| {lender} | {lt} | {in_excel} | {in_db} | {status} |")
    lines.append("")

    # ── Section 9: Excel Sheet Row Counts for Audit Trail
    lines.append("---\n")
    lines.append("## 9. Excel Sheet Audit Summary\n")
    lines.append("| Sheet | File | Rows Read | Lenders Found |")
    lines.append("|-------|------|-----------|---------------|")

    elig_lenders = set(normalize_lender(r.get("Lender_Name")) for r in excel_elig_rows if r.get("Lender_Name"))
    foir_lenders = set(r["lender"] for r in excel_foir_rows if r.get("lender"))
    pf_lenders = set(r["lender"] for r in excel_pf_rows if r.get("lender"))
    login_lenders = set(r["lender"] for r in excel_login_rows if r.get("lender"))

    lines.append(f"| Eligibility Workbook | eligibility workbook (1).xlsx | {len(excel_elig_rows)} | {len(elig_lenders)} |")
    lines.append(f"| FOIR Sheet | FOIR_Sheet (1).xlsx | {len(excel_foir_rows)} | {len(foir_lenders)} |")
    lines.append(f"| PF Data | PF_data (1).xlsx | {len(excel_pf_rows)} | {len(pf_lenders)} |")
    lines.append(f"| Login Fees | Login_fees (1).xlsx | {len(excel_login_rows)} | {len(login_lenders)} |")
    lines.append(f"| HL LTV | HL_LTV_Sheet.xlsx | {len(excel_hl_ltv_rows)} | — |")
    lines.append(f"| LAP LTV | LAP_LTV_Sheet.xlsx | {len(excel_lap_ltv_rows)} | — |")
    lines.append("")

    return "\n".join(lines)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("═" * 70)
    print("  PRYME Engine — Excel ↔ Database Reconciliation")
    print("═" * 70)

    # ── Step 1: Verify all Excel files exist
    print("\n[1/6] Checking Excel files...")
    for key, path in EXCEL_FILES.items():
        if path.exists():
            print(f"  ✅ {key}: {path.name}")
        else:
            print(f"  ❌ {key}: NOT FOUND at {path}")

    # ── Step 2: Read Excel sheets
    print("\n[2/6] Reading Excel sheets...")

    excel_elig_rows = []
    if EXCEL_FILES["eligibility"].exists():
        excel_elig_rows = read_excel_sheet(EXCEL_FILES["eligibility"])
        print(f"  Eligibility: {len(excel_elig_rows)} rows")

    excel_foir_raw = []
    if EXCEL_FILES["foir"].exists():
        excel_foir_raw = read_excel_sheet(EXCEL_FILES["foir"])
        print(f"  FOIR: {len(excel_foir_raw)} rows")

    excel_pf_raw = []
    if EXCEL_FILES["pf"].exists():
        excel_pf_raw = read_excel_sheet(EXCEL_FILES["pf"])
        print(f"  PF: {len(excel_pf_raw)} rows")

    excel_login_raw = []
    if EXCEL_FILES["login_fees"].exists():
        excel_login_raw = read_excel_sheet(EXCEL_FILES["login_fees"])
        print(f"  Login Fees: {len(excel_login_raw)} rows")

    excel_hl_ltv_raw = []
    if EXCEL_FILES["hl_ltv"].exists():
        sheets = read_all_sheets(EXCEL_FILES["hl_ltv"])
        for name, rows in sheets.items():
            excel_hl_ltv_raw.extend(rows)
        print(f"  HL LTV: {len(excel_hl_ltv_raw)} rows across {len(sheets)} sheets")

    excel_lap_ltv_raw = []
    if EXCEL_FILES["lap_ltv"].exists():
        excel_lap_ltv_raw = read_excel_sheet(EXCEL_FILES["lap_ltv"])
        print(f"  LAP LTV: {len(excel_lap_ltv_raw)} rows")

    # ── Step 3: Normalize Excel data
    print("\n[3/6] Normalizing Excel data...")
    excel_foir_rows = normalize_foir_excel(excel_foir_raw)
    excel_pf_rows = normalize_pf_excel(excel_pf_raw)
    excel_login_rows = normalize_login_fees_excel(excel_login_raw)
    excel_hl_ltv_rows = normalize_ltv_excel(excel_hl_ltv_raw, "HL") if excel_hl_ltv_raw else []
    excel_lap_ltv_rows = normalize_ltv_excel(excel_lap_ltv_raw, "LAP")
    print(f"  FOIR: {len(excel_foir_rows)} normalized rows")
    print(f"  PF: {len(excel_pf_rows)} normalized rows")
    print(f"  Login: {len(excel_login_rows)} normalized rows")
    print(f"  HL LTV: {len(excel_hl_ltv_rows)} normalized rows")
    print(f"  LAP LTV: {len(excel_lap_ltv_rows)} normalized rows")

    # ── Step 4: Parse DB state from SQL migrations
    print("\n[4/6] Parsing DB state from SQL migrations...")
    db_products = build_db_products()
    print(f"  DB Products: {len(db_products)} product codes")
    for code in sorted(db_products.keys()):
        p = db_products[code]
        print(f"    {code}: {p['lender_name']} {p['loan_type']}")

    # ── Step 5: Run comparisons
    print("\n[5/6] Running comparisons...")

    # 5a: Product coverage
    excel_products_set = set()
    for row in excel_elig_rows:
        lender = normalize_lender(row.get("Lender_Name"))
        lt = safe_str(row.get("Product_Name", "")).upper().strip()
        if lt in ("HL", "HOME LOAN"):
            lt = "HL"
        elif lt in ("LAP", "LOAN AGAINST PROPERTY"):
            lt = "LAP"
        if lender and lt:
            excel_products_set.add((lender, lt))

    db_products_set = set()
    for p in db_products.values():
        db_products_set.add((normalize_lender(p["lender_name"]), p["loan_type"]))

    missing, extra, common = compare_products(excel_elig_rows, db_products)
    print(f"  Missing from DB: {len(missing)}")
    print(f"  Extra in DB: {len(extra)}")
    print(f"  Matched: {len(common)}")

    # 5b: Eligibility attributes
    elig_discrepancies = compare_eligibility_attributes(excel_elig_rows, db_products)
    print(f"  Eligibility discrepancies: {len(elig_discrepancies)}")

    # 5c: FOIR
    foir_discrepancies = compare_foir(
        excel_foir_rows,
        SQL_DIR / "V28__ingest_foir_slabs_existing_schema.sql"
    )
    print(f"  FOIR discrepancies: {len(foir_discrepancies)}")

    # 5d: Login fees
    login_fee_discrepancies = compare_login_fees(excel_login_rows, db_products)
    print(f"  Login fee discrepancies: {len(login_fee_discrepancies)}")

    # 5e: Processing fees
    pf_discrepancies = compare_pf(
        excel_pf_rows,
        SQL_DIR / "V29__create_and_seed_pf_matrix.sql"
    )
    print(f"  PF discrepancies: {len(pf_discrepancies)}")

    # ── Step 6: Generate report
    print("\n[6/6] Generating report...")
    report = generate_report(
        excel_products_set, db_products_set, missing, extra, common,
        elig_discrepancies, foir_discrepancies, login_fee_discrepancies,
        pf_discrepancies, excel_elig_rows, excel_foir_rows, excel_pf_rows,
        excel_login_rows, excel_hl_ltv_rows, excel_lap_ltv_rows,
        db_products
    )

    OUTPUT.write_text(report)
    print(f"\n{'═' * 70}")
    print(f"  ✅ Report written to: {OUTPUT}")
    print(f"  Total discrepancies: {len(elig_discrepancies) + len(foir_discrepancies) + len(login_fee_discrepancies) + len(pf_discrepancies)}")
    print(f"  Missing products: {len(missing)}")
    print(f"  Extra products: {len(extra)}")
    print(f"{'═' * 70}")


if __name__ == "__main__":
    main()
